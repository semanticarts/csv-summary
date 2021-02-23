# CSV Summary Tool

import csv
import re
from argparse import ArgumentParser, FileType
from collections import defaultdict
from os.path import splitext
from operator import itemgetter

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def parse_arguments():
    parser = ArgumentParser()
    parser.add_argument("input", type=FileType("r", encoding="utf-8", errors="replace"),
                        help="CSV or XSL input containing data. "
                             "If XLS/XLSX, will add two extra sheets with summary/sample data. Will update "
                             "in place unless -o is specified. First sheet is treated as the data unless --sheet-name "
                             "is specified.")
    parser.add_argument("-o", "--output_path",
                        help="Path to output .xlsx, default is same directory/name as CSV input",
                        type=FileType("w"))
    parser.add_argument("--category-threshold", type=int, default=100,
                        help="Columns that have equal or less unique values will be treated as a category, "
                             "and their values will be counted and output in the summary.")
    parser.add_argument("--date-format", action="store",
                        default="\\d{1,2}[-/]\\d{1,2}[-/]\\d{2}(\\d{2})?",
                        help="Regular expression for detecting date/time columns.")
    parser.add_argument("--date-time-format", action="store",
                        default="\\d{1,2}[-/]\\d{1,2}[-/]\\d{2}(\\d{2})?[- ]\\d{2}:\\d{2}:\\d{2}(\\.\\d{1,6})?",
                        help="Regular expression for detecting date/time columns.")
    parser.add_argument("-i", "--ignore-value", action="append", default=[],
                        help="Ignore these values from summary, use for blank equivalents such as '?' and 'N/A'")
    parser.add_argument("--sheet-name", action="store",
                        help="Process XLS with the input in the specified sheet. "
                             "If -o is not specified, will add new sheets to input XLS.")
    parser.add_argument("-s", "--num-samples", type=int, default=3,
                        help="Number of rows to sample in transposed view")

    return parser.parse_args()


def summarize_csv():
    args = parse_arguments()
    output_path = args.output_path

    (base, ext) = splitext(args.input.name)

    date_regex = re.compile(args.date_format)
    date_time_regex = re.compile(args.date_time_format)
    num_rows = 0
    headers = None
    all_dates = defaultdict(lambda: True)
    all_date_times = defaultdict(lambda: True)

    if ext.startswith('.xls') or args.sheet_name:
        wb = load_workbook(args.input.name)
        by_column = None
        if not output_path:
            output_path = args.input.name
        if args.sheet_name:
            csv_copy = wb.get_sheet_by_name(args.sheet_name)
        else:
            csv_copy = wb.worksheets[0]
        for row in csv_copy.iter_rows():
            if not headers:
                headers = list(str(cell.value) for cell in row)
                by_column = dict((header, defaultdict(int)) for header in headers)
                continue
            for col, cell in zip(headers, row):
                value = str(cell.value)
                if value in args.ignore_value:
                    cell.value = ""
                else:
                    by_column[col][value] += 1
                    all_dates[col] &= date_regex.fullmatch(value) is not None
                    all_date_times[col] &= date_time_regex.fullmatch(value) is not None
    else:
        if not output_path:
            output_path = base + ".xlsx"

        wb = Workbook()

        reader = csv.reader(args.input, )
        headers = next(reader)
        by_column = dict((header, defaultdict(int)) for header in headers)

        csv_copy = wb.active
        csv_copy.title = "Data"
        csv_copy.append(headers)

        for line in reader:
            csv_copy.append(list("" if val in args.ignore_value else val for val in line))
            for col, value in zip(headers, line):
                if value not in args.ignore_value:
                    by_column[col][value] += 1
                    all_dates[col] &= date_regex.fullmatch(value) is not None
                    all_date_times[col] &= date_time_regex.fullmatch(value) is not None

            num_rows += 1

        header_row(csv_copy)
        auto_width(csv_copy)

    summary = wb.create_sheet("Summary")
    summary.append(headers)
    for index, header in enumerate(headers):
        values = by_column[header]
        if len(values) == num_rows:
            # All values are unique
            summary.cell(row=2, column=index + 1).value = "Unique"
        elif len(values) > 0 and all_date_times[header]:
            # All values are date times
            summary.cell(row=2, column=index + 1).value = "Date/Times"
        elif len(values) > 0 and all_dates[header]:
            # All values are dates
            summary.cell(row=2, column=index + 1).value = "Dates"
        elif len(values) <= args.category_threshold:
            for row, usage in enumerate(
                    f"{k} [{v}]" for k, v in sorted(values.items(), reverse=True, key=itemgetter(1, 0))):
                summary.cell(row=row + 2, column=index + 1).value = usage
    header_row(summary)
    auto_width(summary)

    samples = wb.create_sheet("Samples")
    for row, hdr in enumerate(headers):
        samples.cell(row=row + 1, column=1).value = hdr
    for sample in range(args.num_samples):
        for index in range(len(headers)):
            samples.cell(row=index + 1, column=sample + 2).value = csv_copy.cell(row=sample + 2, column=index + 1).value
    header_col(samples)
    auto_width(samples)

    wb.save(output_path)


def auto_width(sheet):
    column_widths = [0] * len(next(sheet.iter_rows()))
    for row in sheet.iter_rows():
        for i, cell in enumerate(row):
            column_widths[i] = max(column_widths[i], len(str(cell.value)))
    for i, column_width in enumerate(column_widths):
        sheet.column_dimensions[get_column_letter(i + 1)].width = column_width


def header_row(sheet):
    bold = Font(bold=True)
    for col in range(len(next(sheet.iter_rows()))):
        sheet.cell(column=col + 1, row=1).font = bold
    sheet.freeze_panes = 'A2'


def header_col(sheet):
    bold = Font(bold=True)
    for row in range(len(sheet['A'])):
        sheet.cell(column=1, row=row + 1).font = bold
    sheet.freeze_panes = 'B1'


if __name__ == '__main__':
    summarize_csv()
